import os
from openpyxl import load_workbook
from django.conf import settings
from easypark.models import ParkingLocation
from django.core.management.base import BaseCommand
from datetime import datetime
from django.contrib.auth import get_user_model

User = get_user_model()

class Command(BaseCommand):
    help = 'Load data from Excel into ParkingLocation and CustomUser models'

    def handle(self, *args, **options):
        file_path = os.path.join(settings.BASE_DIR, 'easypark', 'fixtures', 'mockup-data.xlsx')

        # โหลดไฟล์ Excel
        wb = load_workbook(file_path)

        # โหลดข้อมูลจากแผ่นข้อมูลของ ParkingLocation
        location_sheet = wb['ParkingLocation']
        for row in location_sheet.iter_rows(min_row=2, values_only=True):
            row = row[:11]  # ดึงเฉพาะ 11 คอลัมน์แรก
            location_id, name, slug, description, total_spots, available_spots, created_at, updated_at, owner_id, camera_url, image = row

            created_at = created_at or datetime.now()
            updated_at = updated_at or datetime.now()

            # ตรวจสอบว่า owner_id มีหรือไม่
            if not owner_id:
                self.stdout.write(self.style.ERROR(f"Skipping {name} because owner_id is missing"))
                continue

            try:
                # ถ้า owner_id ไม่มีในฐานข้อมูล จะทำการสร้างใหม่
                owner = User.objects.get(id=owner_id)
            except User.DoesNotExist:
                # ถ้าไม่พบ user ด้วย id ที่กำหนดไว้ ให้สร้างใหม่
                self.stdout.write(self.style.WARNING(f"Owner with ID {owner_id} does not exist. Creating new user..."))
                owner = User.objects.create(
                    id=owner_id,
                    username=f"user_{owner_id}",
                    role='user',  # ให้ role เป็น 'manager' ตามความต้องการ
                    first_name=name,  # ใช้ชื่อจาก ParkingLocation สำหรับชื่อผู้ใช้
                    email=f"user{owner_id}@example.com",  # เพิ่ม email สมมติ
                )

            # # ตรวจสอบว่า owner เป็น manager หรือไม่
            # if not owner.is_manager():
            #     self.stdout.write(self.style.ERROR(f"Owner with ID {owner_id} is not a manager. Skipping {name}..."))
            #     continue

            # เพิ่มข้อมูลใน ParkingLocation
            location, created = ParkingLocation.objects.get_or_create(
                id=location_id,
                defaults={
                    'name': name,
                    'slug': slug,
                    'description': description,
                    'total_spots': total_spots,
                    'available_spots': available_spots,
                    'created_at': created_at,
                    'updated_at': updated_at,
                    'owner': owner,
                    'camera_url': camera_url,
                    'image': image,
                }
            )

            if created:
                self.stdout.write(f"Created parking location: {name}")
            else:
                self.stdout.write(f"Parking location already exists: {name}")

        self.stdout.write(self.style.SUCCESS("Data loaded successfully!"))
